from openpyxl import Workbook
from pathlib import Path

def create_test_excel():
    """Create a test Excel file with multiple sheets and various data types."""
    # Create workbook with sample data
    wb = Workbook()
    
    # Get the active worksheet
    ws1 = wb.active
    if ws1 is None:
        ws1 = wb.create_sheet("Sheet1")
    else:
        ws1.title = "Sheet1"

    # Define test data
    sheet1_data = [
        ["Name", "Age", "Notes"],
        ["John Doe", "30", "Regular customer"],
        ["Jane Smith", "25", "VIP, priority service"]
    ]

    # Add data to Sheet1
    for row_idx, row_data in enumerate(sheet1_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    # Create and populate Sheet2
    ws2 = wb.create_sheet("Sheet2")
    sheet2_data = [
        ["Product", "Price"],
        ["Widget", "99.99"],
        ["Gadget", "149.99"]
    ]

    # Add data to Sheet2
    for row_idx, row_data in enumerate(sheet2_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    output_path = Path("sample.xlsx")
    wb.save(output_path)
    print(f"Created sample Excel file at: {output_path}")

if __name__ == "__main__":
    create_test_excel()
